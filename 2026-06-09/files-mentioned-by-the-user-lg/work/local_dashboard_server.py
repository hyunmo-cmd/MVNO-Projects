import io
import json
import mimetypes
import re
import sys
import tempfile
import urllib.parse
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUTPUTS = ROOT / "outputs"
HTML = OUTPUTS / "lg-plan-benefit-dashboard.html"
PORT = 8765


def load_msoffcrypto():
    candidates = [
        ROOT / "work" / "serverdeps",
        Path(r"C:\Users\Public\Documents\ESTsoft\CreatorTemp\mvno_deps2"),
        ROOT / "work" / "pydeps",
    ]
    for candidate in candidates:
        if candidate.exists():
            sys.path.insert(0, str(candidate))
            try:
                import msoffcrypto

                return msoffcrypto
            except Exception:
                try:
                    sys.path.remove(str(candidate))
                except ValueError:
                    pass
    raise RuntimeError("암호화 엑셀 해제 도구가 아직 설치되어 있지 않습니다.")


def parse_multipart(body, content_type):
    boundary_match = re.search(r"boundary=(.+)$", content_type)
    if not boundary_match:
        raise ValueError("업로드 형식을 읽지 못했습니다.")
    boundary = boundary_match.group(1).strip().strip('"').encode()
    parts = {}
    for chunk in body.split(b"--" + boundary):
        chunk = chunk.strip(b"\r\n")
        if not chunk or chunk == b"--":
            continue
        header_blob, _, data = chunk.partition(b"\r\n\r\n")
        headers = header_blob.decode("utf-8", errors="replace")
        name_match = re.search(r'name="([^"]+)"', headers)
        filename_match = re.search(r'filename="([^"]*)"', headers)
        if not name_match:
            continue
        name = name_match.group(1)
        parts[name] = {
            "filename": filename_match.group(1) if filename_match else "",
            "data": data.rstrip(b"\r\n"),
            "headers": headers,
        }
    return parts


def cell_text(value):
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d %H:%M")
    return str(value)


def read_workbook(upload_bytes, password):
    try:
        return load_workbook(io.BytesIO(upload_bytes), read_only=True, data_only=True)
    except Exception:
        if not password:
            raise ValueError("암호화된 파일입니다. 파일 암호를 입력해 주세요.")
        msoffcrypto = load_msoffcrypto()
        decrypted = io.BytesIO()
        office = msoffcrypto.OfficeFile(io.BytesIO(upload_bytes))
        office.load_key(password=password)
        office.decrypt(decrypted)
        decrypted.seek(0)
        return load_workbook(decrypted, read_only=True, data_only=True)


def workbook_to_payload(workbook):
    ws = workbook.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("첫 번째 시트가 비어 있습니다.")
    headers = [cell_text(v).strip() for v in rows[0]]
    index = {name: i for i, name in enumerate(headers)}
    required = ["개통번호", "고객명", "요금제코드"]
    missing = [name for name in required if name not in index]
    if missing:
        raise ValueError("필수 컬럼이 없습니다: " + ", ".join(missing))

    def pick(row, name):
        i = index.get(name, -1)
        return cell_text(row[i]) if 0 <= i < len(row) else ""

    customers = []
    for row in rows[1:]:
        if not row or not any(row):
            continue
        plan_code = pick(row, "요금제코드")
        if not plan_code:
            continue
        customers.append(
            {
                "phone": pick(row, "개통번호"),
                "name": pick(row, "고객명"),
                "planCode": plan_code,
                "planName": pick(row, "요금제명"),
                "openedAt": pick(row, "개통일"),
                "status": pick(row, "상태"),
                "joinType": pick(row, "개통구분"),
                "agency": pick(row, "대리점"),
                "channel": pick(row, "총판"),
            }
        )
    return {"customers": customers}


class Handler(BaseHTTPRequestHandler):
    def end_json(self, status, payload):
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(data)

    def do_GET(self):
        path = urllib.parse.urlparse(self.path).path
        if path in ("/", "/dashboard"):
            target = HTML
        else:
            target = (OUTPUTS / path.lstrip("/")).resolve()
            if OUTPUTS.resolve() not in target.parents and target != OUTPUTS.resolve():
                self.send_error(403)
                return
        if not target.exists() or not target.is_file():
            self.send_error(404)
            return
        data = target.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", mimetypes.guess_type(str(target))[0] or "application/octet-stream")
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(data)

    def do_POST(self):
        if urllib.parse.urlparse(self.path).path != "/api/parse-excel":
            self.send_error(404)
            return
        try:
            length = int(self.headers.get("Content-Length", "0"))
            body = self.rfile.read(length)
            parts = parse_multipart(body, self.headers.get("Content-Type", ""))
            upload = parts.get("file")
            password = parts.get("password", {}).get("data", b"").decode("utf-8", errors="ignore")
            if not upload or not upload["data"]:
                raise ValueError("파일이 첨부되지 않았습니다.")
            workbook = read_workbook(upload["data"], password)
            payload = workbook_to_payload(workbook)
            self.end_json(200, payload)
        except Exception as exc:
            self.end_json(400, {"error": str(exc)})

    def log_message(self, format, *args):
        return


if __name__ == "__main__":
    server = ThreadingHTTPServer(("127.0.0.1", PORT), Handler)
    print(f"http://127.0.0.1:{PORT}/dashboard", flush=True)
    server.serve_forever()
